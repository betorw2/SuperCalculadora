# -*- coding: utf-8 -*-
import json, math, os
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJ="/sessions/great-gallant-planck/mnt/claude orcamento"
os.chdir(PROJ)

# ---------- MOTOR (espelho da calculadora) ----------
MODELOS=json.load(open("M2_Biblioteca_Modelos_BOM.json"))["modelos"]
GRAM={"E":345,"B":425,"EB":615}; CART_G={"E":190,"B":230,"EB":230}
FOLHA_UTIL=1.09*0.60
C=dict(corr_kg=7.60,cart_kg=9.00,perda=1.03,tinta_usd_l=79,tinta_ml_m2=6,plast_m2=0.90,cola_g_m2=6,cola_kg=30,
 caixa=12.63,clip_par=0.55,mo=1.46,noz=dict(dep=(18000000-4500000)/7/12,manut=20000,op=3000,aj=2000,enc=0.70,m_min=50,boca=1.40,min_h=1))

def geo(mk):
    ac=pe=ap=cc=ent=0
    for p in MODELOS[mk]["pecas"]:
        nm,q,o,imp,L,A=p; a=L*A/10000*q*C["perda"]
        ac+=a; pe+=a*GRAM[o]/1000; ap+=a*(2 if imp=="4/4" else 1); cc+=a*CART_G[o]/1000*C["cart_kg"]; ent+=(2 if imp=="4/4" else 1)
    return dict(ac=ac,pe=pe,ap=ap,cc=cc,ent=ent,nprat=MODELOS[mk]["nprat"])

def custo(mk,qtd,o):
    g=geo(mk); acopla=o["impr"] in("nozomi","offset")
    c_corr=g["pe"]*C["corr_kg"]; c_cart=g["cc"] if acopla else 0
    c_tinta=g["ap"]*C["tinta_ml_m2"]/1000*C["tinta_usd_l"]*o["dolar"] if o["impr"] in("nozomi","direta") else 0
    c_plast=g["ap"]*C["plast_m2"] if (o["acab"]!="sem" and o["impr"]!="sem") else 0
    c_cola=g["ac"]*C["cola_g_m2"]/1000*C["cola_kg"]; c_clip=g["nprat"]*1.5*C["clip_par"]
    off_var=off_fixo=0
    if o["impr"]=="offset":
        folhas=g["ap"]/(FOLHA_UTIL*(o["off_aprov"]/100))
        off_var=folhas*o["off_milheiro"]/1000
        off_fixo=o["off_cores"]*o["off_chapa"]*g["ent"]+o["off_acerto"]*g["ent"]+o["off_arte"]+o["off_perda"]*g["ent"]*o["off_milheiro"]/1000
    var_u=c_corr+c_cart+c_tinta+c_plast+c_cola+c_clip+C["mo"]+C["caixa"]+off_var
    hm=22*8*3; chn=(C["noz"]["dep"]+C["noz"]["manut"]+(C["noz"]["op"]+C["noz"]["aj"])*(1+C["noz"]["enc"]))/hm
    c_noz=0
    if o["impr"] in("nozomi","direta"):
        th=g["ap"]*qtd/(C["noz"]["m_min"]*C["noz"]["boca"]*60); c_noz=max(th,C["noz"]["min_h"])*chn
    c_faca=o["faca_valor"] if o["faca"]=="nova" else 0
    ci=var_u*qtd+c_noz+off_fixo+c_faca
    comp=[("Corrugado EB/E (Paraibuna)","Matéria-prima",c_corr*qtd),("Papel cartão 230g/190g","Matéria-prima",c_cart*qtd),
      ("Tinta UV Nozomi","Impressão",c_tinta*qtd),("Offset (impressão)","Impressão",off_var*qtd),("Plastificação","Acabamento",c_plast*qtd),
      ("Cola hotmelt","Colagem",c_cola*qtd),("Clips de prateleira","Acessórios",c_clip*qtd),("Mão de obra (acopl.+corte)","Produção",C["mo"]*qtd),
      ("Caixa de transporte","Embalagem",C["caixa"]*qtd),("Hora-máquina Nozomi (lote)","Impressão",c_noz),
      ("Offset chapas+acerto+arte (lote)","Impressão",off_fixo),("Faca (ferramental, lote)","Ferramental",c_faca)]
    comp=[c for c in comp if c[2]>0.005]
    return dict(ci=ci,ci_u=ci/qtd,comp=comp,g=g,var_u=var_u)

def frete(o,qtd):
    if o["frete"]=="direto": tot=o.get("frete_valor",0)
    elif o["frete"]=="rota":
        v=math.ceil(qtd/max(1,o.get("cap",1))); tot=v*(2*o.get("km",0)*o.get("rs_km",0))
    else: tot=0
    return tot,(tot/qtd if qtd else 0)

# ---------- CONFIG do pedido ----------
CFG=dict(mk="slim4", qtd=158, dolar=5.80, impr="nozomi", acab="sem", faca="existente", faca_valor=0,
  frete="retira", lucro=25, comissao=2, imposto=9, taxa_fin=2, prazo=28,
  off_aprov=82, off_chapa=300, off_cores=4, off_acerto=600, off_milheiro=1200, off_perda=150, off_arte=500,
  cliente="HENRIQUE", orc="79158")
fin=CFG["taxa_fin"]/100*CFG["prazo"]/30
div=max(0.05,1-(CFG["lucro"]/100+CFG["comissao"]/100+CFG["imposto"]/100+fin))
r=custo(CFG["mk"],CFG["qtd"],CFG)
ft,fu=frete(CFG,CFG["qtd"])
preco_retira=r["ci_u"]/div
faixas=[100,158,250,500,1000,2000]
faixa_rows=[(q, custo(CFG["mk"],q,CFG)["ci_u"]/div) for q in faixas]
# comparativo impressão (158 un)
cmp_modes=[("nozomi","Nozomi digital"),("offset","Offset"),("sem","Sem impressão")]
cmp_rows=[]
for mode,lbl in cmp_modes:
    rr=custo(CFG["mk"],CFG["qtd"],{**CFG,"impr":mode})
    cmp_rows.append((lbl, rr["ci_u"], rr["ci_u"]/div))

NOME=MODELOS[CFG["mk"]]["nome"]; DIM=MODELOS[CFG["mk"]]["dim_cm"]
def brl(v): return "R$ "+f"{v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

# ===================== PDFs =====================
CORAL=colors.HexColor("#D85A30"); CORALD=colors.HexColor("#993C1D"); LIGHT=colors.HexColor("#FAECE7"); GRAY=colors.HexColor("#888780"); LINE=colors.HexColor("#E0DCD3")
ss=getSampleStyleSheet()
body=ParagraphStyle('body',parent=ss['Normal'],fontName='Helvetica',fontSize=9.5,leading=14)
small=ParagraphStyle('small',parent=ss['Normal'],fontName='Helvetica',fontSize=8,textColor=GRAY,leading=11)
sech=ParagraphStyle('sech',parent=ss['Normal'],fontName='Helvetica-Bold',fontSize=11,textColor=CORALD,spaceBefore=8,spaceAfter=6)
cellL=ParagraphStyle('cellL',parent=ss['Normal'],fontName='Helvetica',fontSize=8.5,leading=11)

def header(st,W):
    h=Table([[Paragraph("<b>MAVIMIX ADESIVOS DECORATIVOS LTDA</b><br/><font size=8 color='#888780'>Avenida Brasil, 12025 — Braz de Pina — Rio de Janeiro/RJ<br/>CNPJ 06.340.575/0001-30 · IE 79247480 · IM 3542360<br/>Fone (21) 3866-9555 · www.m2flex.com.br · fin@m2flex.com.br</font>",body),
        Image("m2_logo.png",width=38*mm,height=38*mm*86/227)]],colWidths=[W*0.68,W*0.32])
    h.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('ALIGN',(1,0),(1,0),'RIGHT')]))
    st+=[h,Spacer(1,4),HRFlowable(width="100%",thickness=2,color=CORAL),Spacer(1,8)]

def comum(st,W):
    st.append(Paragraph("Atendendo à sua solicitação, apresentamos nossa proposta para confecção do display abaixo descrito:",body)); st.append(Spacer(1,8))
    info=Table([[Paragraph(f"<b>À</b>  {CFG['cliente']}",body),Paragraph(f"<b>Orçamento</b> {CFG['orc']}",body)],
      [Paragraph("A/C: Henrique",small),Paragraph("Data: 12/06/2026 · Validade: 15 dias",small)],
      [Paragraph("CNPJ/CPF: (a preencher) · Tel.: (a preencher)",small),Paragraph(f"Pagamento: FATURADO {CFG['prazo']} dias",small)]],colWidths=[W*0.6,W*0.4])
    info.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('BOTTOMPADDING',(0,0),(-1,-1),2),('TOPPADDING',(0,0),(-1,-1),1)]))
    st+=[info,Spacer(1,8)]
    desc=Paragraph(f"<b><font color='#993C1D'>{NOME.upper()}</font></b><br/><br/>"
      f"Display de chão (FSDU) {DIM['larg']} × {DIM['prof']} × {DIM['alt']} cm, <b>{MODELOS[CFG['mk']]['nprat']} prateleiras</b>.<br/>"
      "Estrutura em corrugado onda <b>EB/E</b> (Paraibuna), cartão <b>230g/190g</b> contracolado, impressão digital <b>UV Nozomi</b> (corpo 4/4, demais 4/0), corte-vinco e montagem automática. Caixa de transporte inclusa.<br/>"
      "<font size=8 color='#888780'>patente BR 112020016883-1</font>",body)
    prod=Table([[Image("display_real.png",width=33*mm,height=33*mm*2828/1029),desc]],colWidths=[40*mm,W-40*mm])
    prod.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(1,0),(1,0),10)]))
    st+=[prod,Spacer(1,10)]

def build(mode,outfile):
    doc=SimpleDocTemplate(outfile,pagesize=A4,leftMargin=16*mm,rightMargin=16*mm,topMargin=14*mm,bottomMargin=14*mm)
    W=A4[0]-32*mm; st=[]; header(st,W); comum(st,W)
    st.append(Paragraph(f"Proposta comercial — quantidade {CFG['qtd']} unidades",sech))
    rows=[['Descrição','Preço unit.','Total'],
      [Paragraph("<b>CLIENTE RETIRA NA FÁBRICA</b> (sem frete)",cellL),brl(preco_retira),brl(preco_retira*CFG['qtd'])]]
    it=Table(rows,colWidths=[W-32*mm-32*mm,32*mm,32*mm])
    sty=[('BACKGROUND',(0,0),(-1,0),CORAL),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
      ('FONTSIZE',(0,0),(-1,-1),9),('ALIGN',(-2,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
      ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,LIGHT]),('GRID',(0,0),(-1,-1),0.5,LINE),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]
    it.setStyle(TableStyle(sty)); st+=[it,Spacer(1,4),Paragraph("Imposto ISS incluso (nota de serviço).",small),Spacer(1,12)]
    st.append(Paragraph("Preço por quantidade",sech))
    if mode=="cliente":
        qt=[['Quantidade','Preço unitário','Preço total']]
        for q,pu in faixa_rows: qt.append([f"{q:,}".replace(",","."),brl(pu),brl(pu*q)])
        qcw=[W*0.34,W*0.33,W*0.33]
    else:
        qt=[['Quantidade','Preço unitário','Preço total','vs. 100 un']]; base=faixa_rows[0][1]
        for q,pu in faixa_rows:
            eco='—' if q==faixa_rows[0][0] else f"-{round((1-pu/base)*100)}%"; qt.append([f"{q:,}".replace(",","."),brl(pu),brl(pu*q),eco])
        qcw=[W*0.25,W*0.25,W*0.3,W*0.2]
    qtb=Table(qt,colWidths=qcw); s2=[('BACKGROUND',(0,0),(-1,0),colors.HexColor("#F1EFE8")),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
      ('FONTSIZE',(0,0),(-1,-1),8.5),('ALIGN',(1,0),(-1,-1),'RIGHT'),('GRID',(0,0),(-1,-1),0.5,LINE),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]
    if mode!="cliente": s2.append(('TEXTCOLOR',(3,1),(3,-1),colors.HexColor("#3B6D11")))
    qtb.setStyle(TableStyle(s2)); st+=[qtb,Spacer(1,8)]
    if mode=="interno":
        # comparativo de impressão
        st.append(Paragraph(f"Comparativo de impressão ({CFG['qtd']} un) — USO INTERNO",sech))
        cm=[['Tipo de impressão','Custo unit.','Preço venda unit.']]
        for lbl,ciu,pv in cmp_rows: cm.append([lbl,brl(ciu),brl(pv)])
        cmt=Table(cm,colWidths=[W*0.4,W*0.3,W*0.3])
        cmt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor("#F1EFE8")),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
          ('FONTSIZE',(0,0),(-1,-1),8.5),('ALIGN',(1,0),(-1,-1),'RIGHT'),('GRID',(0,0),(-1,-1),0.5,LINE),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
        st+=[cmt,Spacer(1,10)]
        st.append(Paragraph(f"Memorial de cálculo — custo industrial ({CFG['qtd']} un) — USO INTERNO",sech))
        mem=[['Item de custo','Etapa','Custo total','% do custo']]; tot=sum(c[2] for c in r["comp"])
        for nome,et,v in r["comp"]: mem.append([nome,et,brl(v),f"{v/tot*100:.1f}%"])
        mem.append(['CUSTO INDUSTRIAL TOTAL','',brl(tot),'100%'])
        mt=Table(mem,colWidths=[W*0.40,W*0.22,W*0.23,W*0.15])
        mt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),CORAL),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
          ('FONTSIZE',(0,0),(-1,-1),8.5),('ALIGN',(2,0),(-1,-1),'RIGHT'),('GRID',(0,0),(-1,-1),0.5,LINE),
          ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white,colors.HexColor("#FAF7F2")]),('BACKGROUND',(0,-1),(-1,-1),LIGHT),
          ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('TEXTCOLOR',(0,-1),(-1,-1),CORALD),('TOPPADDING',(0,0),(-1,-1),4.5),('BOTTOMPADDING',(0,0),(-1,-1),4.5)]))
        st+=[mt,Spacer(1,8),Paragraph("Formação de preço",sech),
          Paragraph(f"Custo industrial unitário: <b>{brl(r['ci_u'])}</b><br/>"
            f"Financeiro pró-rata = {CFG['taxa_fin']}% × {CFG['prazo']}/30 = <b>{fin*100:.2f}%</b><br/>"
            f"Divisor markup = 1 − (lucro {CFG['lucro']}% + comissão {CFG['comissao']}% + imposto {CFG['imposto']}% + financeiro {fin*100:.2f}%) = <b>{div:.4f}</b><br/>"
            f"Preço retira = {brl(r['ci_u'])} ÷ {div:.4f} = <b>{brl(preco_retira)}/un</b>",body),Spacer(1,12)]
    st.append(Paragraph("Condições gerais",sech))
    st.append(Paragraph("1. Valores sujeitos a alteração mediante análise dos arquivos finais.  2. Reservamo-nos o direito de entregar 5% a mais ou a menos da quantidade, pelo mesmo valor unitário.  3. Clientes interestaduais sem inscrição estadual: valor final pode variar pelo diferencial de alíquotas.  4. Matérias-primas sujeitas a disponibilidade após confirmação do pedido.  5. Tinta Nozomi cotada em dólar (US$ ref. R$ 5,80) — reajuste conforme câmbio.  6. Faca/ferramental já existente.  7. Prazo de produção a combinar após aprovação de arte.",small))
    st.append(Spacer(1,10)); st.append(HRFlowable(width="100%",thickness=1,color=LINE)); st.append(Spacer(1,4))
    st.append(Paragraph("USO INTERNO — não enviar ao cliente (contém composição de custo)." if mode=="interno" else "Documento comercial — M2 Flex.",small))
    doc.build(st); print("gerado",outfile)

build("cliente","M2_Proposta_Henrique_Slim_158un.pdf")
build("interno","M2_Folha_Calculo_Henrique_Slim_158un.pdf")

# ===================== EXCEL descritivo =====================
FONT="Arial"
wb=Workbook(); hf=PatternFill("solid",fgColor="D85A30"); hfont=Font(name=FONT,bold=True,color="FFFFFF")
thin=Side(style="thin",color="D9D9D9"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
money='R$ #,##0.00'
def style_hdr(ws,row,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(row,c); cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=bd

# Aba Resumo
ws=wb.active; ws.title="Resumo"; ws.sheet_view.showGridLines=False
ws["A1"]="Orçamento — "+NOME+" · Cliente: "+CFG["cliente"]; ws["A1"].font=Font(name=FONT,bold=True,size=14,color="993C1D")
ws.merge_cells("A1:C1")
resumo=[("Modelo",NOME),("Dimensões (cm)",f"{DIM['larg']} × {DIM['prof']} × {DIM['alt']}"),("Prateleiras",MODELOS[CFG['mk']]['nprat']),
 ("Quantidade",CFG["qtd"]),("Impressão","Nozomi digital"),("Faca/ferramental","Já existente (R$ 0)"),
 ("Entrega","Cliente retira na fábrica (sem frete)"),("Dólar ref.","R$ 5,80"),
 ("Lucro","25%"),("Comissão","2%"),("Imposto","9%"),("Financeiro pró-rata",f"{fin*100:.2f}% ({CFG['taxa_fin']}% × {CFG['prazo']}/30)"),
 ("Divisor markup",f"{div:.4f}"),
 ("Custo industrial unit.",r["ci_u"]),("Custo industrial total",r["ci"]),
 ("PREÇO DE VENDA UNIT.",preco_retira),("PREÇO DE VENDA TOTAL",preco_retira*CFG["qtd"])]
rr=3
for k,v in resumo:
    a=ws.cell(rr,1,k); b=ws.cell(rr,2,v)
    a.font=Font(name=FONT,bold=True,size=10); b.font=Font(name=FONT,size=10)
    if isinstance(v,(int,float)) and k not in("Prateleiras","Quantidade"): b.number_format=money
    if k.startswith("PREÇO"):
        a.fill=PatternFill("solid",fgColor="FAECE7"); b.fill=PatternFill("solid",fgColor="FAECE7")
        a.font=Font(name=FONT,bold=True,size=11,color="993C1D"); b.font=Font(name=FONT,bold=True,size=11,color="993C1D")
    rr+=1
ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=34

# Aba Memorial de custo
ws2=wb.create_sheet("Memorial de custo"); ws2.sheet_view.showGridLines=False
ws2.append(["Item de custo","Etapa","Custo unit. (R$)","Custo total (R$)","% do custo"]); style_hdr(ws2,1,5)
tot=sum(c[2] for c in r["comp"])
ri=2
for nome,et,v in r["comp"]:
    ws2.append([nome,et,v/CFG["qtd"],v,v/tot]);
    ws2.cell(ri,3).number_format=money; ws2.cell(ri,4).number_format=money; ws2.cell(ri,5).number_format='0.0%'
    ri+=1
ws2.append(["CUSTO INDUSTRIAL TOTAL","",r["ci_u"],tot,1])
for c in range(1,6):
    ws2.cell(ri,c).font=Font(name=FONT,bold=True,color="993C1D"); ws2.cell(ri,c).fill=PatternFill("solid",fgColor="FAECE7")
ws2.cell(ri,3).number_format=money; ws2.cell(ri,4).number_format=money; ws2.cell(ri,5).number_format='0.0%'
for col,w in zip("ABCDE",[34,16,18,18,12]): ws2.column_dimensions[col].width=w
ws2.freeze_panes="A2"

# Aba BOM (peças)
ws3=wb.create_sheet("BOM (peças)"); ws3.sheet_view.showGridLines=False
ws3.append(["Peça","Qtd","Onda","Impressão","Medida L×A (cm)","Área c/ perda (m²)"]); style_hdr(ws3,1,6)
ri=2
for p in MODELOS[CFG["mk"]]["pecas"]:
    nm,q,o,imp,L,A=p; area=L*A/10000*q*C["perda"]
    ws3.append([nm,q,o,imp,f"{L}×{A}",round(area,3)]); ri+=1
for col,w in zip("ABCDEF",[18,8,10,12,18,18]): ws3.column_dimensions[col].width=w
ws3.freeze_panes="A2"

# Aba Comparativo impressão
ws4=wb.create_sheet("Comparativo impressão"); ws4.sheet_view.showGridLines=False
ws4.append(["Tipo de impressão","Custo industrial unit. (R$)","Preço de venda unit. (R$)","Preço total 158 un (R$)"]); style_hdr(ws4,1,4)
ri=2
for lbl,ciu,pv in cmp_rows:
    ws4.append([lbl,ciu,pv,pv*CFG["qtd"]])
    for cc in (2,3,4): ws4.cell(ri,cc).number_format=money
    if lbl=="Nozomi digital":
        for cc in range(1,5): ws4.cell(ri,cc).fill=PatternFill("solid",fgColor="FAECE7"); ws4.cell(ri,cc).font=Font(name=FONT,bold=True)
    ri+=1
for col,w in zip("ABCD",[22,24,24,22]): ws4.column_dimensions[col].width=w
ws4.freeze_panes="A2"

# Aba Preço por quantidade
ws5=wb.create_sheet("Preço por quantidade"); ws5.sheet_view.showGridLines=False
ws5.append(["Quantidade","Preço unitário (R$)","Preço total (R$)","Economia vs. 100 un"]); style_hdr(ws5,1,4)
base=faixa_rows[0][1]; ri=2
for q,pu in faixa_rows:
    eco=0 if q==faixa_rows[0][0] else (1-pu/base)
    ws5.append([q,pu,pu*q,eco]); ws5.cell(ri,2).number_format=money; ws5.cell(ri,3).number_format=money; ws5.cell(ri,4).number_format='0%'
    if q==CFG["qtd"]:
        for cc in range(1,5): ws5.cell(ri,cc).fill=PatternFill("solid",fgColor="FAECE7"); ws5.cell(ri,cc).font=Font(name=FONT,bold=True)
    ri+=1
for col,w in zip("ABCD",[14,20,20,20]): ws5.column_dimensions[col].width=w
ws5.freeze_panes="A2"

xlsx_out="M2_Orcamento_Henrique_Slim_158un.xlsx"
wb.save(xlsx_out); print("gerado",xlsx_out)
print(f"\nci_u={r['ci_u']:.2f}  div={div:.4f}  preco_retira={preco_retira:.2f}  total={preco_retira*CFG['qtd']:.2f}")
